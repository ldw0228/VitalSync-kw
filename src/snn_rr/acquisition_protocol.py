"""Offline reconstruction of the seven-stage HAI acquisition protocol.

The acquisition spreadsheet contains lossless operator notes for every
session and hand-written intervals for S02/S03.  This module reads that source
without modifying it, separates the phase-7 assignment from whole-session
metadata, and supplies an ordered dynamic-programming decoder for the sessions
without manual intervals.

This is an annotation and quality-control layer, not a deployable feature
extractor.  In particular, BIOPAC-derived marker annotations are explicitly
forbidden as model inference features.  They may only support offline
alignment, label construction, or retrospective stage evaluation.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
import hashlib
import io
from pathlib import Path
import math
import os
import re
import stat
from typing import Any, Iterable, Mapping, Sequence

import numpy as np
from openpyxl import load_workbook
import yaml


SCHEMA_VERSION = "acquisition_protocol_v1"
STAGE_IDS = tuple(f"phase{index}" for index in range(1, 8))

ANNOTATION_USAGE_CONTRACT: dict[str, Any] = {
    "purpose": "offline_acquisition_segmentation_and_evaluation",
    "inference_feature_allowed": False,
    "biopac_derived_annotation": {
        "inference_feature_allowed": False,
        "permitted_uses": (
            "offline_alignment_audit",
            "label_construction",
            "retrospective_stage_evaluation",
        ),
    },
    "phase7_assignment": {
        "inference_feature_allowed": False,
        "semantics": "assigned action in phase 7 only; not a whole-session activity label",
    },
}


class ProtocolConfigError(ValueError):
    """Raised when an acquisition-protocol configuration is invalid."""


class ProtocolDecodeError(ValueError):
    """Raised when ordered boundaries cannot satisfy manual timing anchors."""


def _source_snapshot(
    path: str | Path, *, label: str, expected_sha256: str | None = None
) -> tuple[bytes, str]:
    """Capture the exact regular-file bytes parsed by protocol consumers."""

    source = Path(path).expanduser().absolute()
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    descriptor = -1
    try:
        before_path = os.stat(source, follow_symlinks=False)
        descriptor = os.open(source, flags)
        before_fd = os.fstat(descriptor)
        if not (
            stat.S_ISREG(before_path.st_mode)
            and stat.S_ISREG(before_fd.st_mode)
            and before_path.st_nlink == before_fd.st_nlink == 1
            and (before_path.st_dev, before_path.st_ino)
            == (before_fd.st_dev, before_fd.st_ino)
        ):
            raise ValueError(f"{label} must be an unaliased regular file: {source}")
        chunks: list[bytes] = []
        digest = hashlib.sha256()
        while True:
            chunk = os.read(descriptor, 1024 * 1024)
            if not chunk:
                break
            chunks.append(chunk)
            digest.update(chunk)
        payload = b"".join(chunks)
        after_fd = os.fstat(descriptor)
        after_path = os.stat(source, follow_symlinks=False)
    except OSError as error:
        raise ValueError(f"cannot snapshot {label}: {source} ({error})") from error
    finally:
        if descriptor >= 0:
            os.close(descriptor)
    signature = lambda value: (
        value.st_dev,
        value.st_ino,
        value.st_mode,
        value.st_nlink,
        value.st_size,
        value.st_mtime_ns,
        value.st_ctime_ns,
    )
    if not (
        signature(before_fd) == signature(after_fd) == signature(after_path)
        and before_fd.st_size == len(payload)
    ):
        raise ValueError(f"{label} changed while being snapshotted: {source}")
    observed_sha256 = digest.hexdigest()
    if expected_sha256 is not None and observed_sha256 != expected_sha256:
        raise ValueError(
            f"{label} consumed-byte SHA-256 mismatch: "
            f"expected {expected_sha256}, observed {observed_sha256}"
        )
    return payload, observed_sha256


@dataclass(frozen=True, slots=True)
class StageSpec:
    stage_id: str
    order: int
    name: str
    nominal_duration_s: float
    duration_scale_s: float
    plausible_duration_s: tuple[float, float]
    nominal_gap_before_s: float
    gap_scale_s: float
    plausible_gap_before_s: tuple[float, float]


@dataclass(frozen=True, slots=True)
class DecoderSpec:
    candidate_grid_step_s: float
    marker_match_sigma_s: float
    duration_penalty_weight: float
    gap_penalty_weight: float
    plausible_range_penalty_weight: float
    manual_boundaries_are_locked: bool
    minimum_stage_duration_s: float
    auto_threshold: float
    review_threshold: float
    manual_boundary_confidence: float
    prior_only_boundary_confidence: float
    source_weights: Mapping[str, float]


@dataclass(frozen=True, slots=True)
class WindowAssignmentSpec:
    minimum_overlap_fraction: float
    transition_guard_s: float


@dataclass(frozen=True, slots=True)
class AcquisitionProtocolConfig:
    schema_version: str
    time_basis: str
    stages: tuple[StageSpec, ...]
    decoder: DecoderSpec
    window_assignment: WindowAssignmentSpec
    v2_candidate_from_session: int
    fall_v1_id: str
    fall_v2_id: str
    annotation_contract: Mapping[str, Any]


@dataclass(frozen=True, slots=True)
class ManualInterval:
    """One slash-separated timing entry preserved from the spreadsheet."""

    start_s: float | None
    end_s: float | None
    raw: str
    complete: bool
    source: str = "dataset_issue_spreadsheet"
    qc_flags: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class IssueAnnotation:
    code: str
    severity: str
    stage_ids: tuple[str, ...]
    evidence: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class FallProtocolCandidate:
    protocol_id: str
    confidence: float
    rationale: str


@dataclass(frozen=True, slots=True)
class SessionProtocolRecord:
    session_id: str
    session_number: int
    acquisition_date: str | None
    participant_name: str | None
    reported_subject_label: str | None
    folder_subject_label: str | None
    physical_identity: str | None
    height_cm: float | None
    weight_kg: float | None
    age_years: int | None
    phase7_assignment: str | None
    issue_text: str | None
    extra_notes: str | None
    issues: tuple[IssueAnnotation, ...]
    qc_flags: tuple[str, ...]
    manual_intervals: Mapping[str, tuple[ManualInterval, ...]]
    fall_protocol_candidates: tuple[FallProtocolCandidate, ...]
    annotation_schema_version: str = SCHEMA_VERSION
    annotation_inference_feature_allowed: bool = False

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["manual_intervals"] = {
            stage_id: [interval.to_dict() for interval in intervals]
            for stage_id, intervals in self.manual_intervals.items()
        }
        result["issues"] = [issue.to_dict() for issue in self.issues]
        result["fall_protocol_candidates"] = [
            asdict(candidate) for candidate in self.fall_protocol_candidates
        ]
        return result


@dataclass(frozen=True, slots=True)
class BoundaryCandidate:
    """Offline marker evidence supplied to the ordered decoder.

    ``boundary`` is ``"start"``, ``"end"``, or ``None``.  ``stage_id`` may
    also be omitted for an unlabeled marker.  Set ``biopac_derived`` whenever
    any part of the candidate was computed from RSP/BIOPAC; the resulting
    annotation remains forbidden as an inference feature.
    """

    time_s: float
    score: float = 1.0
    source: str = "unknown"
    stage_id: str | None = None
    boundary: str | None = None
    biopac_derived: bool = False


@dataclass(frozen=True, slots=True)
class DecodedBoundary:
    time_s: float
    kind: str
    confidence: float
    status: str
    source: str
    evidence_sources: tuple[str, ...] = ()
    manual: bool = False


@dataclass(frozen=True, slots=True)
class StageAttempt:
    """A preserved timing entry; slash-separated entries are not assumed retries."""

    attempt_index: int
    start_s: float | None
    end_s: float | None
    complete: bool
    source: str
    interpretation: str
    qc_flags: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class DecodedStage:
    stage_id: str
    name: str
    start: DecodedBoundary
    end: DecodedBoundary
    duration_s: float
    confidence: float
    status: str
    attempts: tuple[StageAttempt, ...]
    qc_flags: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class DecodedProtocol:
    session_id: str | None
    duration_s: float
    stages: tuple[DecodedStage, ...]
    status: str
    confidence: float
    qc_flags: tuple[str, ...]
    phase7_assignment: str | None
    path_score: float
    path_margin: float | None
    used_biopac_derived_annotations: bool
    annotation_schema_version: str = SCHEMA_VERSION
    annotation_inference_feature_allowed: bool = False
    biopac_annotation_inference_feature_allowed: bool = False

    def stage(self, stage_id: str) -> DecodedStage:
        for item in self.stages:
            if item.stage_id == stage_id:
                return item
        raise KeyError(stage_id)

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["stages"] = [stage.to_dict() for stage in self.stages]
        result["annotation_contract"] = ANNOTATION_USAGE_CONTRACT
        return result


@dataclass(frozen=True, slots=True)
class WindowStageAssignment:
    window_start_s: float
    window_end_s: float
    stage_id: str | None
    overlap_fraction: float
    transition_guard_triggered: bool
    eligible_for_stage_metrics: bool
    reason: str
    stage_confidence: float | None
    phase7_assignment: str | None


def _require_exact_keys(
    value: Mapping[Any, Any], expected: set[str], label: str
) -> None:
    """Reject missing, unknown, or mistyped keys in a versioned contract."""

    actual = set(value)
    if actual != expected:
        missing = sorted(expected - actual)
        unknown = sorted((repr(item) for item in actual - expected))
        raise ProtocolConfigError(
            f"{label} keys mismatch; missing={missing}, unknown={unknown}"
        )


def _as_yaml_number(value: Any, field_name: str) -> float:
    """Return one finite YAML number without coercing strings or booleans."""

    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ProtocolConfigError(f"{field_name} must be a finite YAML number")
    result = float(value)
    if not math.isfinite(result):
        raise ProtocolConfigError(f"{field_name} must be finite")
    return result


def _as_float_pair(value: Any, field_name: str) -> tuple[float, float]:
    if (
        not isinstance(value, Sequence)
        or isinstance(value, (str, bytes))
        or len(value) != 2
    ):
        raise ProtocolConfigError(f"{field_name} must contain exactly two numbers")
    pair = (
        _as_yaml_number(value[0], f"{field_name}[0]"),
        _as_yaml_number(value[1], f"{field_name}[1]"),
    )
    if not 0 <= pair[0] <= pair[1]:
        raise ProtocolConfigError(f"invalid {field_name}: {pair}")
    return pair


def load_protocol_config(
    path: str | Path, *, expected_sha256: str | None = None
) -> AcquisitionProtocolConfig:
    """Load and validate the versioned seven-stage protocol contract."""

    config_path = Path(path)
    try:
        payload, _ = _source_snapshot(
            config_path,
            label="protocol configuration",
            expected_sha256=expected_sha256,
        )
        document = yaml.safe_load(payload.decode("utf-8"))
    except (UnicodeError, yaml.YAMLError, ValueError) as error:
        raise ProtocolConfigError(
            f"cannot load protocol configuration: {error}"
        ) from error
    if not isinstance(document, Mapping):
        raise ProtocolConfigError("protocol config must be a mapping")
    _require_exact_keys(
        document,
        {
            "schema_version",
            "time_basis",
            "annotation_contract",
            "stages",
            "fall_protocol",
            "decoder",
            "window_assignment",
        },
        "protocol config",
    )
    if document.get("schema_version") != SCHEMA_VERSION:
        raise ProtocolConfigError(f"expected schema_version={SCHEMA_VERSION!r}")
    if document.get("time_basis") != "seconds_from_biopac_start":
        raise ProtocolConfigError(
            "time_basis must equal 'seconds_from_biopac_start'"
        )

    stage_documents = document.get("stages")
    if not isinstance(stage_documents, list) or len(stage_documents) != 7:
        raise ProtocolConfigError("exactly seven stages are required")
    stages: list[StageSpec] = []
    for item in stage_documents:
        if not isinstance(item, Mapping):
            raise ProtocolConfigError("each stage must be a mapping")
        _require_exact_keys(
            item,
            {
                "id",
                "order",
                "name",
                "nominal_duration_s",
                "duration_scale_s",
                "plausible_duration_s",
                "nominal_gap_before_s",
                "gap_scale_s",
                "plausible_gap_before_s",
            },
            "stage",
        )
        stage_id = item.get("id")
        stage_name = item.get("name")
        if not isinstance(stage_id, str) or not stage_id:
            raise ProtocolConfigError("stage id must be a non-empty YAML string")
        if not isinstance(stage_name, str) or not stage_name:
            raise ProtocolConfigError("stage name must be a non-empty YAML string")
        raw_order = item.get("order")
        if type(raw_order) is not int:
            raise ProtocolConfigError("stage order must be a YAML integer")
        stage = StageSpec(
            stage_id=stage_id,
            order=raw_order,
            name=stage_name,
            nominal_duration_s=_as_yaml_number(
                item.get("nominal_duration_s"), "nominal_duration_s"
            ),
            duration_scale_s=_as_yaml_number(
                item.get("duration_scale_s"), "duration_scale_s"
            ),
            plausible_duration_s=_as_float_pair(
                item.get("plausible_duration_s"), "plausible_duration_s"
            ),
            nominal_gap_before_s=_as_yaml_number(
                item.get("nominal_gap_before_s"), "nominal_gap_before_s"
            ),
            gap_scale_s=_as_yaml_number(
                item.get("gap_scale_s"), "gap_scale_s"
            ),
            plausible_gap_before_s=_as_float_pair(
                item.get("plausible_gap_before_s"), "plausible_gap_before_s"
            ),
        )
        if (
            not math.isfinite(stage.nominal_duration_s)
            or not math.isfinite(stage.duration_scale_s)
            or stage.nominal_duration_s <= 0
            or stage.duration_scale_s <= 0
        ):
            raise ProtocolConfigError(f"invalid duration prior for {stage.stage_id}")
        if (
            not math.isfinite(stage.nominal_gap_before_s)
            or not math.isfinite(stage.gap_scale_s)
            or stage.nominal_gap_before_s < 0
            or stage.gap_scale_s <= 0
        ):
            raise ProtocolConfigError(f"invalid gap prior for {stage.stage_id}")
        stages.append(stage)
    stages.sort(key=lambda stage: stage.order)
    if tuple(stage.stage_id for stage in stages) != STAGE_IDS:
        raise ProtocolConfigError(f"stage order must be {STAGE_IDS}")
    if tuple(stage.order for stage in stages) != tuple(range(1, 8)):
        raise ProtocolConfigError("stage order values must be 1 through 7")

    decoder_doc = document.get("decoder")
    if not isinstance(decoder_doc, Mapping):
        raise ProtocolConfigError("decoder must be a mapping")
    _require_exact_keys(
        decoder_doc,
        {
            "candidate_grid_step_s",
            "marker_match_sigma_s",
            "duration_penalty_weight",
            "gap_penalty_weight",
            "plausible_range_penalty_weight",
            "manual_boundaries_are_locked",
            "minimum_stage_duration_s",
            "confidence",
            "source_weights",
        },
        "decoder",
    )
    confidence_doc = decoder_doc.get("confidence")
    if not isinstance(confidence_doc, Mapping):
        raise ProtocolConfigError("decoder.confidence must be a mapping")
    _require_exact_keys(
        confidence_doc,
        {
            "auto_threshold",
            "review_threshold",
            "manual_boundary",
            "prior_only_boundary",
        },
        "decoder.confidence",
    )
    source_weights_doc = decoder_doc.get("source_weights")
    if not isinstance(source_weights_doc, Mapping) or not source_weights_doc:
        raise ProtocolConfigError("decoder.source_weights must be a non-empty mapping")
    _require_exact_keys(
        source_weights_doc,
        {
            "radar_marker",
            "radar_motion",
            "biopac_marker",
            "fused_marker",
            "unknown",
        },
        "decoder.source_weights",
    )
    source_weights: dict[str, float] = {}
    for key, value in source_weights_doc.items():
        if not isinstance(key, str) or not key:
            raise ProtocolConfigError(
                "decoder.source_weights keys must be non-empty YAML strings"
            )
        source_weights[key] = _as_yaml_number(
            value, f"decoder.source_weights.{key}"
        )
    manual_locked = decoder_doc.get("manual_boundaries_are_locked")
    if type(manual_locked) is not bool:
        raise ProtocolConfigError(
            "manual_boundaries_are_locked must be a YAML boolean"
        )
    decoder = DecoderSpec(
        candidate_grid_step_s=_as_yaml_number(
            decoder_doc.get("candidate_grid_step_s"),
            "decoder.candidate_grid_step_s",
        ),
        marker_match_sigma_s=_as_yaml_number(
            decoder_doc.get("marker_match_sigma_s"),
            "decoder.marker_match_sigma_s",
        ),
        duration_penalty_weight=_as_yaml_number(
            decoder_doc.get("duration_penalty_weight"),
            "decoder.duration_penalty_weight",
        ),
        gap_penalty_weight=_as_yaml_number(
            decoder_doc.get("gap_penalty_weight"),
            "decoder.gap_penalty_weight",
        ),
        plausible_range_penalty_weight=_as_yaml_number(
            decoder_doc.get("plausible_range_penalty_weight"),
            "decoder.plausible_range_penalty_weight",
        ),
        manual_boundaries_are_locked=manual_locked,
        minimum_stage_duration_s=_as_yaml_number(
            decoder_doc.get("minimum_stage_duration_s"),
            "decoder.minimum_stage_duration_s",
        ),
        auto_threshold=_as_yaml_number(
            confidence_doc.get("auto_threshold"),
            "decoder.confidence.auto_threshold",
        ),
        review_threshold=_as_yaml_number(
            confidence_doc.get("review_threshold"),
            "decoder.confidence.review_threshold",
        ),
        manual_boundary_confidence=_as_yaml_number(
            confidence_doc.get("manual_boundary"),
            "decoder.confidence.manual_boundary",
        ),
        prior_only_boundary_confidence=_as_yaml_number(
            confidence_doc.get("prior_only_boundary"),
            "decoder.confidence.prior_only_boundary",
        ),
        source_weights=source_weights,
    )
    decoder_numeric = (
        decoder.candidate_grid_step_s,
        decoder.marker_match_sigma_s,
        decoder.duration_penalty_weight,
        decoder.gap_penalty_weight,
        decoder.plausible_range_penalty_weight,
        decoder.minimum_stage_duration_s,
        decoder.auto_threshold,
        decoder.review_threshold,
        decoder.manual_boundary_confidence,
        decoder.prior_only_boundary_confidence,
        *decoder.source_weights.values(),
    )
    if not all(math.isfinite(value) for value in decoder_numeric):
        raise ProtocolConfigError("decoder numeric settings must be finite")
    if decoder.candidate_grid_step_s <= 0 or decoder.marker_match_sigma_s <= 0:
        raise ProtocolConfigError("decoder grid and marker sigma must be positive")
    if (
        decoder.duration_penalty_weight < 0
        or decoder.gap_penalty_weight < 0
        or decoder.plausible_range_penalty_weight < 0
        or decoder.minimum_stage_duration_s <= 0
        or any(value < 0 for value in decoder.source_weights.values())
    ):
        raise ProtocolConfigError("decoder penalties/weights are outside valid ranges")
    if not 0 <= decoder.review_threshold <= decoder.auto_threshold <= 1:
        raise ProtocolConfigError("confidence thresholds must satisfy 0 <= review <= auto <= 1")
    if not (
        0 <= decoder.manual_boundary_confidence <= 1
        and 0 <= decoder.prior_only_boundary_confidence <= 1
    ):
        raise ProtocolConfigError("boundary confidence values must be in [0,1]")

    window_doc = document.get("window_assignment")
    if not isinstance(window_doc, Mapping):
        raise ProtocolConfigError("window_assignment must be a mapping")
    _require_exact_keys(
        window_doc,
        {"minimum_overlap_fraction", "transition_guard_s"},
        "window_assignment",
    )
    window_assignment = WindowAssignmentSpec(
        minimum_overlap_fraction=_as_yaml_number(
            window_doc.get("minimum_overlap_fraction"),
            "window_assignment.minimum_overlap_fraction",
        ),
        transition_guard_s=_as_yaml_number(
            window_doc.get("transition_guard_s"),
            "window_assignment.transition_guard_s",
        ),
    )
    if not math.isfinite(window_assignment.minimum_overlap_fraction) or not (
        0 < window_assignment.minimum_overlap_fraction <= 1
    ):
        raise ProtocolConfigError("minimum_overlap_fraction must be in (0, 1]")
    if (
        not math.isfinite(window_assignment.transition_guard_s)
        or window_assignment.transition_guard_s < 0
    ):
        raise ProtocolConfigError("transition_guard_s must be nonnegative")

    fall_doc = document.get("fall_protocol")
    if not isinstance(fall_doc, Mapping):
        raise ProtocolConfigError("fall_protocol must be a mapping")
    _require_exact_keys(
        fall_doc,
        {"v2_candidate_from_session", "v1_id", "v2_id"},
        "fall_protocol",
    )
    v2_candidate_from_session = fall_doc.get("v2_candidate_from_session")
    if type(v2_candidate_from_session) is not int or v2_candidate_from_session < 1:
        raise ProtocolConfigError(
            "fall_protocol.v2_candidate_from_session must be a positive YAML integer"
        )
    fall_v1_id = fall_doc.get("v1_id")
    fall_v2_id = fall_doc.get("v2_id")
    if (
        not isinstance(fall_v1_id, str)
        or not fall_v1_id
        or not isinstance(fall_v2_id, str)
        or not fall_v2_id
        or fall_v1_id == fall_v2_id
    ):
        raise ProtocolConfigError(
            "fall protocol IDs must be distinct non-empty YAML strings"
        )
    annotation_contract = document.get("annotation_contract")
    if not isinstance(annotation_contract, Mapping):
        raise ProtocolConfigError("annotation_contract must be a mapping")
    _require_exact_keys(
        annotation_contract,
        {
            "purpose",
            "inference_feature_allowed",
            "biopac_derived_annotation",
            "phase7_assignment",
        },
        "annotation_contract",
    )
    if annotation_contract.get("purpose") != ANNOTATION_USAGE_CONTRACT["purpose"]:
        raise ProtocolConfigError("annotation_contract purpose is not canonical")
    if annotation_contract.get("inference_feature_allowed") is not False:
        raise ProtocolConfigError("offline annotations must be forbidden as inference features")
    biopac_contract = annotation_contract.get("biopac_derived_annotation")
    if not isinstance(biopac_contract, Mapping):
        raise ProtocolConfigError(
            "annotation_contract.biopac_derived_annotation must be a mapping"
        )
    _require_exact_keys(
        biopac_contract,
        {"inference_feature_allowed", "permitted_uses"},
        "annotation_contract.biopac_derived_annotation",
    )
    if biopac_contract.get("inference_feature_allowed") is not False:
        raise ProtocolConfigError("BIOPAC-derived annotations must be forbidden at inference")
    permitted_uses = biopac_contract.get("permitted_uses")
    expected_permitted_uses = ANNOTATION_USAGE_CONTRACT[
        "biopac_derived_annotation"
    ]["permitted_uses"]
    if not isinstance(permitted_uses, list) or tuple(permitted_uses) != tuple(
        expected_permitted_uses
    ):
        raise ProtocolConfigError("BIOPAC annotation permitted uses are not canonical")
    phase7_contract = annotation_contract.get("phase7_assignment")
    if not isinstance(phase7_contract, Mapping):
        raise ProtocolConfigError(
            "annotation_contract.phase7_assignment must be a mapping"
        )
    _require_exact_keys(
        phase7_contract,
        {"inference_feature_allowed", "semantics"},
        "annotation_contract.phase7_assignment",
    )
    if (
        phase7_contract.get("inference_feature_allowed") is not False
        or phase7_contract.get("semantics")
        != ANNOTATION_USAGE_CONTRACT["phase7_assignment"]["semantics"]
    ):
        raise ProtocolConfigError("phase7 assignment annotation is not canonical")
    return AcquisitionProtocolConfig(
        schema_version=SCHEMA_VERSION,
        time_basis="seconds_from_biopac_start",
        stages=tuple(stages),
        decoder=decoder,
        window_assignment=window_assignment,
        v2_candidate_from_session=v2_candidate_from_session,
        fall_v1_id=fall_v1_id,
        fall_v2_id=fall_v2_id,
        annotation_contract=annotation_contract,
    )


_INTERVAL_RE = re.compile(
    r"^\s*(?P<start>\d+(?:\.\d+)?)?\s*[~～]\s*(?P<end>\d+(?:\.\d+)?)?\s*$"
)


def parse_manual_interval_cell(value: Any) -> tuple[ManualInterval, ...]:
    """Parse slash-separated intervals while preserving partial/manual text."""

    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ()
    text = str(value).strip()
    if not text:
        return ()
    intervals: list[ManualInterval] = []
    for raw_part in re.split(r"\s*/\s*", text):
        raw = raw_part.strip()
        if not raw:
            continue
        match = _INTERVAL_RE.fullmatch(raw)
        if match is None:
            intervals.append(
                ManualInterval(
                    start_s=None,
                    end_s=None,
                    raw=raw,
                    complete=False,
                    qc_flags=("manual_interval_unparsed",),
                )
            )
            continue
        start = float(match.group("start")) if match.group("start") else None
        end = float(match.group("end")) if match.group("end") else None
        flags: list[str] = []
        if start is None:
            flags.append("manual_start_missing")
        if end is None:
            flags.append("manual_end_missing")
        if start is not None and end is not None and end <= start:
            flags.append("manual_interval_nonpositive")
        intervals.append(
            ManualInterval(
                start_s=start,
                end_s=end,
                raw=raw,
                complete=start is not None and end is not None and end > start,
                qc_flags=tuple(flags),
            )
        )
    return tuple(intervals)


def _normalise_optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalise_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _normalise_age(value: Any) -> int | None:
    if value is None:
        return None
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else None


def _session_directories(parent: Path) -> dict[int, str]:
    result: dict[int, str] = {}
    if not parent.is_dir():
        return result
    for path in parent.iterdir():
        match = re.fullmatch(r"S(\d{2})_([A-Za-z0-9]+)", path.name)
        if path.is_dir() and match:
            number = int(match.group(1))
            if number in result:
                raise ValueError(f"multiple session directories for S{number:02d}")
            result[number] = path.name
    return result


@dataclass(frozen=True, slots=True)
class _IssueRule:
    code: str
    pattern: re.Pattern[str]
    severity: str
    stages: tuple[str, ...]


_ISSUE_RULES = (
    _IssueRule("biopac_clipping", re.compile(r"clipping|10\s*~\s*-10", re.I), "review", ()),
    _IssueRule(
        "biopac_attachment_concern",
        re.compile(r"바이오팩|BIOPAC|흘러\s*내려", re.I),
        "uncertain",
        (),
    ),
    _IssueRule(
        "uwb_reference_alignment_concern",
        re.compile(r"UWB.*(?:래퍼런스|엇나)|(?:래퍼런스|가이드).*엇나", re.I | re.S),
        "review",
        (),
    ),
    _IssueRule(
        "missing_stage_marker",
        re.compile(r"마커를?\s*(?:주지\s*않|안\s*주)|마커\s*없", re.I),
        "review",
        ("phase4",),
    ),
    _IssueRule("excess_markers", re.compile(r"마커를?\s*많이"), "uncertain", ()),
    _IssueRule(
        "weak_marker_execution",
        re.compile(r"마커.*(?:대충|설렁설렁)"),
        "uncertain",
        (),
    ),
    _IssueRule(
        "marker_motion_contamination",
        re.compile(r"마커.*(?:인형|레이더\s*움직임)|인형.*레이더\s*움직임", re.S),
        "review",
        ("phase3",),
    ),
    _IssueRule(
        "marker_ended_early",
        re.compile(r"마커.*일찍\s*끝"),
        "review",
        ("phase4",),
    ),
    _IssueRule(
        "retry_reported",
        re.compile(r"retry|재촬영|다시\s*측정", re.I),
        "uncertain",
        (),
    ),
    _IssueRule(
        "phase7_may_be_unusable",
        re.compile(r"자유\s*동작.*못\s*쓸"),
        "review",
        ("phase7",),
    ),
    _IssueRule(
        "fall_v2_scenario_added",
        re.compile(r"낙상.*물건\s*집는\s*시나리오\s*추가", re.S),
        "uncertain",
        ("phase4",),
    ),
    _IssueRule(
        "route_or_cornering_deviation",
        re.compile(r"지그재그.*(?:retry|코너링|칸\s*이동)", re.I | re.S),
        "uncertain",
        ("phase5",),
    ),
    _IssueRule(
        "phase2_activity_variant",
        re.compile(r"스쿼트\s*\+\s*팔굽혀펴기"),
        "uncertain",
        ("phase2",),
    ),
    _IssueRule(
        "phase2_orientation_deviation",
        re.compile(r"운동\s*후.*1번\s*레이더", re.S),
        "uncertain",
        ("phase2",),
    ),
    _IssueRule(
        "phase2_wait_omitted",
        re.compile(r"15초\s*대기\s*안"),
        "uncertain",
        ("phase2",),
    ),
    _IssueRule(
        "round_trip_pace_deviation",
        re.compile(r"빨리\s*걸었다.*왕복|왕복.*빨리", re.S),
        "uncertain",
        ("phase6",),
    ),
    _IssueRule(
        "fall_speed_contrast_weak",
        re.compile(r"낙상\s*속도\s*차이|천천히\s*눕기.*철푸덕", re.S),
        "uncertain",
        ("phase4",),
    ),
    _IssueRule(
        "phase7_repetition_deviation",
        re.compile(r"발차기\s*3번"),
        "uncertain",
        ("phase7",),
    ),
    _IssueRule(
        "phase7_action_ambiguous",
        re.compile(r"스트라이크\s*동작\s*애매"),
        "review",
        ("phase7",),
    ),
    _IssueRule(
        "fall_execution_deviation",
        re.compile(r"낙상.*(?:바로\s*안\s*나감|넘어지는거\s*좀\s*느림)", re.S),
        "uncertain",
        ("phase4",),
    ),
    _IssueRule(
        "post_exercise_response_weak",
        re.compile(r"운동\s*후\s*호흡.*빠르지\s*않"),
        "uncertain",
        ("phase2",),
    ),
    _IssueRule(
        "unplanned_interstage_movement",
        re.compile(r"1번\s*실험\s*끝나고.*이동", re.S),
        "uncertain",
        ("phase1", "phase2"),
    ),
)


def _infer_retry_stages(text: str) -> tuple[str, ...]:
    stages: list[str] = []
    keywords = {
        "phase3": ("픽업",),
        "phase4": ("낙상", "넘어"),
        "phase5": ("지그", "코너링", "칸"),
        "phase6": ("왕복",),
        "phase7": ("자유", "스트라이크", "발차기"),
    }
    for stage_id, words in keywords.items():
        if any(word in text for word in words):
            stages.append(stage_id)
    return tuple(stages)


def _structure_issues(issue_text: str | None, extra_notes: str | None) -> tuple[IssueAnnotation, ...]:
    parts = [text for text in (issue_text, extra_notes) if text]
    if not parts:
        return ()
    combined = "\n".join(parts)
    annotations: list[IssueAnnotation] = []
    for rule in _ISSUE_RULES:
        if rule.pattern.search(combined):
            stages = rule.stages
            if rule.code == "retry_reported" and not stages:
                stages = _infer_retry_stages(combined)
            annotations.append(
                IssueAnnotation(
                    code=rule.code,
                    severity=rule.severity,
                    stage_ids=stages,
                    evidence=combined,
                )
            )
    if not annotations:
        annotations.append(
            IssueAnnotation(
                code="reported_issue_unclassified",
                severity="uncertain",
                stage_ids=(),
                evidence=combined,
            )
        )
    return tuple(annotations)


def _fall_candidates(
    session_number: int,
    issues: Sequence[IssueAnnotation],
    config: AcquisitionProtocolConfig,
) -> tuple[FallProtocolCandidate, ...]:
    explicit_v2 = any(issue.code == "fall_v2_scenario_added" for issue in issues)
    if explicit_v2:
        return (
            FallProtocolCandidate(config.fall_v2_id, 0.98, "operator note says added and retained"),
            FallProtocolCandidate(config.fall_v1_id, 0.02, "older guide variant retained as fallback"),
        )
    if session_number >= config.v2_candidate_from_session:
        return (
            FallProtocolCandidate(
                config.fall_v2_id,
                0.80,
                f"session is S{session_number:02d} or later after the S04 protocol change",
            ),
            FallProtocolCandidate(
                config.fall_v1_id,
                0.20,
                "visible and hidden guide slides retain conflicting variants",
            ),
        )
    return (
        FallProtocolCandidate(
            config.fall_v1_id,
            0.80,
            "session predates the S04 note introducing the fourth scenario",
        ),
        FallProtocolCandidate(
            config.fall_v2_id,
            0.20,
            "actual adherence was not recorded per scenario",
        ),
    )


def _physical_identity(session_id: str, reported_label: str | None) -> str | None:
    # The authoritative identity table also resolves S17_RJS -> PJS and the
    # CJW/CHW spreadsheet-folder discrepancy.  Import lazily to keep this
    # parser independent of signal-processing imports during module loading.
    try:
        from .preprocess import SESSION_IDENTITY

        return SESSION_IDENTITY.get(session_id, reported_label)
    except ImportError:
        return reported_label


def _exact_spreadsheet_session_number(value: Any, worksheet_row: int) -> int:
    """Validate the physical-session selector without numeric coercion."""

    if type(value) is not int or value <= 0:
        raise ValueError(
            "acquisition spreadsheet session number must be a positive "
            f"exact integer at worksheet row {worksheet_row}: {value!r}"
        )
    return value


def load_dataset_issue_records(
    workbook_path: str | Path,
    *,
    config: AcquisitionProtocolConfig,
    dataset_root: str | Path | None = None,
    expected_sha256: str | None = None,
) -> tuple[SessionProtocolRecord, ...]:
    """Read all session rows in ``Dataset_issue.xlsx`` without writing it.

    Dates in merged cells are forward-filled.  Raw issue text is retained even
    when a structured QC rule matches it.  If ``dataset_root`` is supplied,
    canonical on-disk session directory names take precedence over worksheet
    label spelling, while the reported label remains preserved separately.
    """

    path = Path(workbook_path)
    root = Path(dataset_root) if dataset_root is not None else path.parent
    session_dirs = _session_directories(root)
    workbook_payload, _ = _source_snapshot(
        path,
        label="dataset issue spreadsheet",
        expected_sha256=expected_sha256,
    )
    workbook = load_workbook(
        io.BytesIO(workbook_payload), read_only=True, data_only=True
    )
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise ValueError(f"empty workbook: {path}") from exc
        stage_columns: dict[int, str] = {}
        for index, value in enumerate(header):
            match = re.match(r"\s*([1-7])번", str(value)) if value is not None else None
            if match:
                stage_columns[index] = f"phase{match.group(1)}"
        if tuple(stage_columns.values()) != STAGE_IDS:
            raise ValueError("spreadsheet must contain ordered 1번 through 7번 columns")

        records: list[SessionProtocolRecord] = []
        current_date: date | datetime | str | None = None
        for worksheet_row, row in enumerate(rows, start=2):
            values = list(row) + [None] * max(0, 17 - len(row))
            session_number_raw = values[3]
            if session_number_raw is None:
                continue
            # Session number selects the physical source directory and thereby
            # its identity, protocol annotations, and later evaluation roles.
            # Excel booleans are integers in Python and ``int(1.5)`` truncates;
            # accepting either would silently attach one participant's row to
            # another session.  Blank rows remain skippable, but any populated
            # non-canonical identifier invalidates the workbook.
            session_number = _exact_spreadsheet_session_number(
                session_number_raw, worksheet_row
            )
            if values[0] is not None:
                current_date = values[0]
            if isinstance(current_date, (date, datetime)):
                date_text = (
                    current_date.date().isoformat()
                    if isinstance(current_date, datetime)
                    else current_date.isoformat()
                )
            else:
                date_text = _normalise_optional_text(current_date)

            reported_label = _normalise_optional_text(values[2])
            session_id = session_dirs.get(session_number)
            if session_id is None:
                suffix = reported_label or "UNKNOWN"
                session_id = f"S{session_number:02d}_{suffix}"
            folder_label = session_id.split("_", maxsplit=1)[1] if "_" in session_id else None
            issue_text = _normalise_optional_text(values[8])
            extra_notes = _normalise_optional_text(values[16])
            issues = list(_structure_issues(issue_text, extra_notes))
            qc_flags = {issue.code for issue in issues}
            if reported_label and folder_label and reported_label != folder_label:
                qc_flags.add("folder_reported_label_mismatch")
                issues.append(
                    IssueAnnotation(
                        code="folder_reported_label_mismatch",
                        severity="review",
                        stage_ids=(),
                        evidence=f"folder={folder_label}; spreadsheet={reported_label}",
                    )
                )

            manual_intervals = {
                stage_id: parse_manual_interval_cell(values[column])
                for column, stage_id in stage_columns.items()
            }
            for intervals in manual_intervals.values():
                for interval in intervals:
                    qc_flags.update(interval.qc_flags)

            record = SessionProtocolRecord(
                session_id=session_id,
                session_number=session_number,
                acquisition_date=date_text,
                participant_name=_normalise_optional_text(values[1]),
                reported_subject_label=reported_label,
                folder_subject_label=folder_label,
                physical_identity=_physical_identity(session_id, reported_label or folder_label),
                height_cm=_normalise_number(values[4]),
                weight_kg=_normalise_number(values[5]),
                age_years=_normalise_age(values[6]),
                phase7_assignment=_normalise_optional_text(values[7]),
                issue_text=issue_text,
                extra_notes=extra_notes,
                issues=tuple(issues),
                qc_flags=tuple(sorted(qc_flags)),
                manual_intervals=manual_intervals,
                fall_protocol_candidates=_fall_candidates(session_number, issues, config),
            )
            records.append(record)
    finally:
        workbook.close()

    records.sort(key=lambda item: item.session_number)
    if len({record.session_number for record in records}) != len(records):
        raise ValueError("duplicate session numbers in acquisition spreadsheet")
    return tuple(records)


def records_by_session(
    records: Iterable[SessionProtocolRecord],
) -> dict[str, SessionProtocolRecord]:
    result: dict[str, SessionProtocolRecord] = {}
    for record in records:
        if record.session_id in result:
            raise ValueError(f"duplicate session_id: {record.session_id}")
        result[record.session_id] = record
    return result


def _manual_anchors(record: SessionProtocolRecord | None) -> dict[tuple[str, str], float]:
    anchors: dict[tuple[str, str], float] = {}
    if record is None:
        return anchors
    for stage_id, intervals in record.manual_intervals.items():
        starts = [interval.start_s for interval in intervals if interval.start_s is not None]
        ends = [interval.end_s for interval in intervals if interval.end_s is not None]
        if starts:
            anchors[(stage_id, "start")] = float(min(starts))
        if ends:
            anchors[(stage_id, "end")] = float(max(ends))
    return anchors


def _boundary_descriptors(
    stages: Sequence[StageSpec],
) -> tuple[tuple[StageSpec, str], ...]:
    return tuple((stage, kind) for stage in stages for kind in ("start", "end"))


def _nominal_boundaries(stages: Sequence[StageSpec]) -> dict[tuple[str, str], float]:
    time_s = 0.0
    result: dict[tuple[str, str], float] = {}
    for stage in stages:
        time_s += stage.nominal_gap_before_s
        result[(stage.stage_id, "start")] = time_s
        time_s += stage.nominal_duration_s
        result[(stage.stage_id, "end")] = time_s
    return result


def _validate_candidates(
    candidates: Iterable[BoundaryCandidate], duration_s: float
) -> tuple[BoundaryCandidate, ...]:
    result: list[BoundaryCandidate] = []
    for candidate in candidates:
        if not math.isfinite(candidate.time_s) or not 0 <= candidate.time_s <= duration_s:
            raise ValueError(f"candidate time outside session: {candidate.time_s}")
        if not math.isfinite(candidate.score) or not 0 <= candidate.score <= 1:
            raise ValueError("candidate score must be finite and in [0, 1]")
        if candidate.stage_id is not None and candidate.stage_id not in STAGE_IDS:
            raise ValueError(f"unknown candidate stage: {candidate.stage_id}")
        if candidate.boundary not in (None, "start", "end"):
            raise ValueError("candidate boundary must be start, end, or None")
        result.append(candidate)
    return tuple(result)


def _huber(value: np.ndarray | float) -> np.ndarray | float:
    absolute = np.abs(value)
    return np.where(absolute <= 1.0, 0.5 * absolute**2, absolute - 0.5)


def _outside_range_penalty(
    value: np.ndarray,
    plausible_range: tuple[float, float],
    scale: float,
) -> np.ndarray:
    below = np.maximum(plausible_range[0] - value, 0.0)
    above = np.maximum(value - plausible_range[1], 0.0)
    return (below + above) / max(scale, np.finfo(float).eps)


def _transition_score(
    delta: np.ndarray,
    *,
    nominal: float,
    scale: float,
    plausible_range: tuple[float, float],
    weight: float,
    range_weight: float,
) -> np.ndarray:
    normalized = (delta - nominal) / scale
    return -weight * np.asarray(_huber(normalized)) - range_weight * _outside_range_penalty(
        delta, plausible_range, scale
    )


def _candidate_compatible(
    candidate: BoundaryCandidate,
    stage_id: str,
    kind: str,
) -> bool:
    return (candidate.stage_id in (None, stage_id)) and (candidate.boundary in (None, kind))


def _emission_scores(
    times: np.ndarray,
    candidates: Sequence[BoundaryCandidate],
    stage_id: str,
    kind: str,
    decoder: DecoderSpec,
) -> np.ndarray:
    result = np.zeros(times.shape, dtype=np.float64)
    sigma = decoder.marker_match_sigma_s
    for candidate in candidates:
        if not _candidate_compatible(candidate, stage_id, kind):
            continue
        weight = decoder.source_weights.get(
            candidate.source, decoder.source_weights.get("unknown", 0.7)
        )
        distance = (times - candidate.time_s) / sigma
        result += weight * candidate.score * np.exp(-0.5 * distance**2)
    return result


def _boundary_evidence(
    time_s: float,
    candidates: Sequence[BoundaryCandidate],
    stage_id: str,
    kind: str,
    decoder: DecoderSpec,
) -> tuple[float, tuple[str, ...]]:
    support = 0.0
    sources: set[str] = set()
    for candidate in candidates:
        if not _candidate_compatible(candidate, stage_id, kind):
            continue
        distance = abs(time_s - candidate.time_s)
        if distance > 3.0 * decoder.marker_match_sigma_s:
            continue
        weight = decoder.source_weights.get(
            candidate.source, decoder.source_weights.get("unknown", 0.7)
        )
        support += weight * candidate.score * math.exp(
            -0.5 * (distance / decoder.marker_match_sigma_s) ** 2
        )
        sources.add(candidate.source)
    return support, tuple(sorted(sources))


def _status_from_confidence(confidence: float, decoder: DecoderSpec) -> str:
    if confidence >= decoder.auto_threshold:
        return "auto"
    if confidence >= decoder.review_threshold:
        return "uncertain"
    return "review"


def _manual_attempts(
    stage_id: str,
    record: SessionProtocolRecord | None,
    decoded_start: float,
    decoded_end: float,
) -> tuple[StageAttempt, ...]:
    intervals = record.manual_intervals.get(stage_id, ()) if record is not None else ()
    if not intervals:
        return (
            StageAttempt(
                attempt_index=1,
                start_s=decoded_start,
                end_s=decoded_end,
                complete=True,
                source="ordered_dp_decoder",
                interpretation="decoded_stage_envelope",
            ),
        )
    return tuple(
        StageAttempt(
            attempt_index=index,
            start_s=interval.start_s,
            end_s=interval.end_s,
            complete=interval.complete,
            source=interval.source,
            interpretation="manual_subinterval_not_necessarily_retry",
            qc_flags=interval.qc_flags,
        )
        for index, interval in enumerate(intervals, start=1)
    )


def _stage_qc_flags(
    stage_id: str,
    record: SessionProtocolRecord | None,
    attempts: Sequence[StageAttempt],
) -> tuple[str, ...]:
    flags = {flag for attempt in attempts for flag in attempt.qc_flags}
    if record is not None:
        for issue in record.issues:
            if not issue.stage_ids or stage_id in issue.stage_ids:
                flags.add(issue.code)
    return tuple(sorted(flags))


def _qc_forces_review(flags: Sequence[str]) -> bool:
    return bool(
        {
            "biopac_clipping",
            "uwb_reference_alignment_concern",
            "missing_stage_marker",
            "marker_motion_contamination",
            "marker_ended_early",
            "phase7_may_be_unusable",
            "phase7_action_ambiguous",
            "manual_interval_unparsed",
            "manual_interval_nonpositive",
        }.intersection(flags)
    )


def _status_with_qc(
    status: str,
    stage_id: str,
    flags: Sequence[str],
    record: SessionProtocolRecord | None,
) -> str:
    if _qc_forces_review(flags):
        return "review"
    if "manual_start_missing" in flags or "manual_end_missing" in flags:
        status = "uncertain" if status == "auto" else status
    if record is None:
        return status
    applicable = [
        issue for issue in record.issues if not issue.stage_ids or stage_id in issue.stage_ids
    ]
    if any(issue.severity == "review" for issue in applicable):
        return "review"
    if status == "auto" and any(issue.severity == "uncertain" for issue in applicable):
        return "uncertain"
    return status


def decode_ordered_protocol(
    *,
    duration_s: float,
    config: AcquisitionProtocolConfig,
    candidates: Iterable[BoundaryCandidate] = (),
    session_record: SessionProtocolRecord | None = None,
) -> DecodedProtocol:
    """Decode fourteen ordered stage boundaries with a Viterbi-style DP.

    Candidate evidence supplies node scores.  Alternating duration and gap
    transitions receive Huber soft penalties relative to the versioned
    protocol priors.  Complete or partial spreadsheet boundaries are inserted
    as exact locked states by default, so automatic marker evidence cannot
    silently replace a human timing entry.
    """

    if not math.isfinite(duration_s) or duration_s <= 0:
        raise ValueError("duration_s must be positive and finite")
    evidence = _validate_candidates(candidates, duration_s)
    anchors = _manual_anchors(session_record)
    nominal = _nominal_boundaries(config.stages)
    grid = np.arange(0.0, duration_s + config.decoder.candidate_grid_step_s, config.decoder.candidate_grid_step_s)
    grid = grid[grid <= duration_s]
    extra_times = [0.0, duration_s]
    extra_times.extend(candidate.time_s for candidate in evidence)
    extra_times.extend(value for value in anchors.values() if 0 <= value <= duration_s)
    extra_times.extend(value for value in nominal.values() if 0 <= value <= duration_s)
    times = np.unique(np.concatenate((grid, np.asarray(extra_times, dtype=np.float64))))
    descriptors = _boundary_descriptors(config.stages)

    score_layers: list[np.ndarray] = []
    parent_layers: list[np.ndarray] = []
    first_stage, first_kind = descriptors[0]
    first_emission = _emission_scores(
        times, evidence, first_stage.stage_id, first_kind, config.decoder
    )
    first_delta = times
    first_scores = first_emission + _transition_score(
        first_delta,
        nominal=first_stage.nominal_gap_before_s,
        scale=first_stage.gap_scale_s,
        plausible_range=first_stage.plausible_gap_before_s,
        weight=config.decoder.gap_penalty_weight,
        range_weight=config.decoder.plausible_range_penalty_weight,
    )
    first_anchor = anchors.get((first_stage.stage_id, first_kind))
    if first_anchor is not None and config.decoder.manual_boundaries_are_locked:
        first_scores[~np.isclose(times, first_anchor, rtol=0.0, atol=1e-9)] = -np.inf
    score_layers.append(first_scores)
    parent_layers.append(np.full(times.shape, -1, dtype=np.int32))

    for position in range(1, len(descriptors)):
        stage, kind = descriptors[position]
        previous_stage, previous_kind = descriptors[position - 1]
        delta = times[:, None] - times[None, :]
        if kind == "end" and previous_stage.stage_id == stage.stage_id:
            valid = delta >= config.decoder.minimum_stage_duration_s
            transition = _transition_score(
                delta,
                nominal=stage.nominal_duration_s,
                scale=stage.duration_scale_s,
                plausible_range=stage.plausible_duration_s,
                weight=config.decoder.duration_penalty_weight,
                range_weight=config.decoder.plausible_range_penalty_weight,
            )
        else:
            valid = delta >= 0.0
            transition = _transition_score(
                delta,
                nominal=stage.nominal_gap_before_s,
                scale=stage.gap_scale_s,
                plausible_range=stage.plausible_gap_before_s,
                weight=config.decoder.gap_penalty_weight,
                range_weight=config.decoder.plausible_range_penalty_weight,
            )
        transition[~valid] = -np.inf
        total = transition + score_layers[-1][None, :]
        parents = np.argmax(total, axis=1).astype(np.int32)
        best = total[np.arange(len(times)), parents]
        emission = _emission_scores(times, evidence, stage.stage_id, kind, config.decoder)
        current_scores = best + emission
        anchor = anchors.get((stage.stage_id, kind))
        if anchor is not None and config.decoder.manual_boundaries_are_locked:
            current_scores[~np.isclose(times, anchor, rtol=0.0, atol=1e-9)] = -np.inf
        score_layers.append(current_scores)
        parent_layers.append(parents)

    final_scores = score_layers[-1]
    finite_final = final_scores[np.isfinite(final_scores)]
    if finite_final.size == 0:
        raise ProtocolDecodeError("manual timing anchors are inconsistent with ordered stages")
    final_index = int(np.argmax(final_scores))
    path_score = float(final_scores[final_index])
    if finite_final.size > 1:
        best_two = np.partition(finite_final, -2)[-2:]
        path_margin = float(best_two.max() - best_two.min())
    else:
        path_margin = None
    path_indices = [final_index]
    for position in range(len(descriptors) - 1, 0, -1):
        path_indices.append(int(parent_layers[position][path_indices[-1]]))
    path_indices.reverse()
    path_times = [float(times[index]) for index in path_indices]

    decoded_stages: list[DecodedStage] = []
    for stage_index, stage in enumerate(config.stages):
        start_time = path_times[2 * stage_index]
        end_time = path_times[2 * stage_index + 1]
        decoded_boundaries: list[DecodedBoundary] = []
        for kind, boundary_time in (("start", start_time), ("end", end_time)):
            manual = (stage.stage_id, kind) in anchors
            support, sources = _boundary_evidence(
                boundary_time, evidence, stage.stage_id, kind, config.decoder
            )
            if manual:
                confidence = config.decoder.manual_boundary_confidence
                source = "dataset_issue_spreadsheet"
            elif support > 0:
                confidence = config.decoder.prior_only_boundary_confidence + (
                    1.0 - config.decoder.prior_only_boundary_confidence
                ) * (1.0 - math.exp(-support))
                source = "marker_evidence"
            else:
                confidence = config.decoder.prior_only_boundary_confidence
                source = "protocol_prior"
            decoded_boundaries.append(
                DecodedBoundary(
                    time_s=boundary_time,
                    kind=kind,
                    confidence=float(np.clip(confidence, 0.0, 1.0)),
                    status=_status_from_confidence(confidence, config.decoder),
                    source=source,
                    evidence_sources=sources,
                    manual=manual,
                )
            )
        duration = end_time - start_time
        duration_deviation = (duration - stage.nominal_duration_s) / stage.duration_scale_s
        duration_plausibility = math.exp(-0.5 * min(duration_deviation**2, 16.0))
        boundary_confidence = math.sqrt(
            decoded_boundaries[0].confidence * decoded_boundaries[1].confidence
        )
        confidence = float(np.clip(boundary_confidence * (0.7 + 0.3 * duration_plausibility), 0, 1))
        attempts = _manual_attempts(stage.stage_id, session_record, start_time, end_time)
        stage_flags = _stage_qc_flags(stage.stage_id, session_record, attempts)
        status = _status_from_confidence(confidence, config.decoder)
        status = _status_with_qc(status, stage.stage_id, stage_flags, session_record)
        decoded_stages.append(
            DecodedStage(
                stage_id=stage.stage_id,
                name=stage.name,
                start=decoded_boundaries[0],
                end=decoded_boundaries[1],
                duration_s=duration,
                confidence=confidence,
                status=status,
                attempts=attempts,
                qc_flags=stage_flags,
            )
        )

    status_rank = {"auto": 0, "uncertain": 1, "review": 2}
    overall_status = max(decoded_stages, key=lambda stage: status_rank[stage.status]).status
    overall_confidence = float(np.mean([stage.confidence for stage in decoded_stages]))
    all_flags = set(session_record.qc_flags if session_record is not None else ())
    all_flags.update(flag for stage in decoded_stages for flag in stage.qc_flags)
    return DecodedProtocol(
        session_id=session_record.session_id if session_record is not None else None,
        duration_s=float(duration_s),
        stages=tuple(decoded_stages),
        status=overall_status,
        confidence=overall_confidence,
        qc_flags=tuple(sorted(all_flags)),
        phase7_assignment=session_record.phase7_assignment if session_record is not None else None,
        path_score=path_score,
        path_margin=path_margin,
        used_biopac_derived_annotations=any(item.biopac_derived for item in evidence),
    )


def assign_window_to_stage(
    window_start_s: float,
    window_end_s: float,
    protocol: DecodedProtocol,
    *,
    minimum_overlap_fraction: float = 0.80,
    transition_guard_s: float = 2.0,
) -> WindowStageAssignment:
    """Assign one half-open window while excluding boundary transitions."""

    if not (
        math.isfinite(window_start_s)
        and math.isfinite(window_end_s)
        and 0 <= window_start_s < window_end_s <= protocol.duration_s
    ):
        raise ValueError("window must be finite, positive, and inside the session")
    if not 0 < minimum_overlap_fraction <= 1:
        raise ValueError("minimum_overlap_fraction must be in (0, 1]")
    if transition_guard_s < 0:
        raise ValueError("transition_guard_s must be nonnegative")

    window_duration = window_end_s - window_start_s
    overlaps: list[tuple[float, DecodedStage]] = []
    guard_triggered = False
    for stage in protocol.stages:
        overlap = max(0.0, min(window_end_s, stage.end.time_s) - max(window_start_s, stage.start.time_s))
        overlaps.append((overlap / window_duration, stage))
        for boundary in (stage.start.time_s, stage.end.time_s):
            if window_start_s - transition_guard_s < boundary < window_end_s + transition_guard_s:
                guard_triggered = True
    overlap_fraction, winner = max(overlaps, key=lambda item: item[0])
    if guard_triggered:
        return WindowStageAssignment(
            window_start_s,
            window_end_s,
            None,
            overlap_fraction,
            True,
            False,
            "transition_guard",
            winner.confidence if overlap_fraction > 0 else None,
            None,
        )
    if overlap_fraction < minimum_overlap_fraction:
        return WindowStageAssignment(
            window_start_s,
            window_end_s,
            None,
            overlap_fraction,
            False,
            False,
            "insufficient_stage_overlap",
            winner.confidence if overlap_fraction > 0 else None,
            None,
        )
    phase7_assignment = protocol.phase7_assignment if winner.stage_id == "phase7" else None
    metric_eligible = winner.status == "auto"
    if metric_eligible:
        reason = "assigned"
    elif winner.status == "uncertain":
        reason = "stage_uncertain"
    elif winner.status == "review":
        reason = "stage_requires_review"
    else:
        # Serialized or externally constructed protocol objects must not make
        # an unknown status eligible by virtue of merely being non-"review".
        reason = "stage_status_invalid"
    return WindowStageAssignment(
        window_start_s,
        window_end_s,
        winner.stage_id,
        overlap_fraction,
        False,
        metric_eligible,
        reason,
        winner.confidence,
        phase7_assignment,
    )


def assign_windows_to_stages(
    windows: Iterable[tuple[float, float]],
    protocol: DecodedProtocol,
    *,
    config: AcquisitionProtocolConfig,
) -> tuple[WindowStageAssignment, ...]:
    spec = config.window_assignment
    return tuple(
        assign_window_to_stage(
            start,
            end,
            protocol,
            minimum_overlap_fraction=spec.minimum_overlap_fraction,
            transition_guard_s=spec.transition_guard_s,
        )
        for start, end in windows
    )
