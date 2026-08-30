import copy
from dataclasses import replace
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from snn_rr.acquisition_protocol import (
    ANNOTATION_USAGE_CONTRACT,
    BoundaryCandidate,
    ProtocolConfigError,
    _exact_spreadsheet_session_number,
    assign_window_to_stage,
    decode_ordered_protocol,
    load_dataset_issue_records,
    load_protocol_config,
    parse_manual_interval_cell,
    records_by_session,
)


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "configs" / "acquisition_protocol_v1.yaml"
WORKBOOK_PATH = ROOT / "HAI_EXPERIMENT" / "Dataset_issue.xlsx"


def _loaded():
    config = load_protocol_config(CONFIG_PATH)
    records = load_dataset_issue_records(
        WORKBOOK_PATH,
        config=config,
        dataset_root=ROOT / "HAI_EXPERIMENT",
    )
    return config, records_by_session(records)


def test_protocol_config_has_seven_soft_stage_priors_and_target_firewall():
    config = load_protocol_config(CONFIG_PATH)
    assert [stage.stage_id for stage in config.stages] == [f"phase{i}" for i in range(1, 8)]
    assert config.stages[1].nominal_duration_s == 360.0
    assert all(stage.duration_scale_s > 0 for stage in config.stages)
    assert all(stage.gap_scale_s > 0 for stage in config.stages)
    assert config.annotation_contract["inference_feature_allowed"] is False
    assert (
        config.annotation_contract["biopac_derived_annotation"]["inference_feature_allowed"]
        is False
    )
    assert ANNOTATION_USAGE_CONTRACT["phase7_assignment"]["inference_feature_allowed"] is False


def test_protocol_sources_bind_the_exact_consumed_bytes() -> None:
    with pytest.raises(ProtocolConfigError, match="consumed-byte SHA-256 mismatch"):
        load_protocol_config(CONFIG_PATH, expected_sha256="0" * 64)
    config = load_protocol_config(CONFIG_PATH)
    with pytest.raises(ValueError, match="consumed-byte SHA-256 mismatch"):
        load_dataset_issue_records(
            WORKBOOK_PATH,
            config=config,
            dataset_root=ROOT / "HAI_EXPERIMENT",
            expected_sha256="0" * 64,
        )


@pytest.mark.parametrize(
    "value", [True, 1.5, float("nan"), float("inf"), float("-inf"), "2", 0, -1]
)
def test_spreadsheet_session_number_rejects_coercive_values(value: object) -> None:
    with pytest.raises(ValueError, match="positive exact integer"):
        _exact_spreadsheet_session_number(value, 2)
    assert _exact_spreadsheet_session_number(1, 2) == 1


def test_workbook_fractional_session_number_cannot_select_physical_session(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(WORKBOOK_PATH)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet.cell(row=2, column=4, value=1.5)
    path = tmp_path / "fractional-session.xlsx"
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="worksheet row 2"):
        load_dataset_issue_records(
            path,
            dataset_root=ROOT / "HAI_EXPERIMENT",
            config=load_protocol_config(CONFIG_PATH),
        )


@pytest.mark.parametrize("value", ["false", 0, 1, float("nan")])
def test_protocol_config_rejects_non_boolean_manual_lock(
    tmp_path: Path, value
) -> None:
    document = copy.deepcopy(yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")))
    document["decoder"]["manual_boundaries_are_locked"] = value
    path = tmp_path / "bad.yaml"
    path.write_text(yaml.safe_dump(document), encoding="utf-8")
    with pytest.raises(ProtocolConfigError, match="YAML boolean"):
        load_protocol_config(path)


def test_protocol_config_rejects_nonfinite_decoder_number(tmp_path: Path) -> None:
    document = copy.deepcopy(yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")))
    document["decoder"]["duration_penalty_weight"] = float("nan")
    path = tmp_path / "bad-number.yaml"
    path.write_text(yaml.safe_dump(document), encoding="utf-8")
    with pytest.raises(ProtocolConfigError, match="must be finite"):
        load_protocol_config(path)


@pytest.mark.parametrize(
    ("path", "value"),
    [
        (("stages", 0, "nominal_duration_s"), True),
        (("stages", 0, "nominal_duration_s"), "180.0"),
        (("stages", 0, "plausible_duration_s", 0), True),
        (("stages", 0, "plausible_gap_before_s", 1), "120.0"),
        (("decoder", "duration_penalty_weight"), True),
        (("decoder", "marker_match_sigma_s"), "1.5"),
        (("decoder", "confidence", "auto_threshold"), "0.80"),
        (("decoder", "source_weights", "radar_marker"), True),
        (("window_assignment", "transition_guard_s"), "2.0"),
        (("fall_protocol", "v2_candidate_from_session"), True),
        (("fall_protocol", "v2_candidate_from_session"), "4"),
    ],
)
def test_protocol_config_rejects_bool_and_string_numeric_coercion(
    tmp_path: Path, path: tuple[object, ...], value: object
) -> None:
    document = copy.deepcopy(yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")))
    target = document
    for key in path[:-1]:
        target = target[key]
    target[path[-1]] = value
    config_path = tmp_path / "coercive-number.yaml"
    config_path.write_text(yaml.safe_dump(document), encoding="utf-8")

    with pytest.raises(ProtocolConfigError):
        load_protocol_config(config_path)


@pytest.mark.parametrize(
    "mapping_path",
    [
        (),
        ("stages", 0),
        ("decoder",),
        ("decoder", "confidence"),
        ("decoder", "source_weights"),
        ("window_assignment",),
        ("fall_protocol",),
        ("annotation_contract",),
        ("annotation_contract", "biopac_derived_annotation"),
        ("annotation_contract", "phase7_assignment"),
    ],
)
@pytest.mark.parametrize("mutation", ["unknown", "missing"])
def test_protocol_config_requires_exact_key_inventory(
    tmp_path: Path, mapping_path: tuple[object, ...], mutation: str
) -> None:
    document = copy.deepcopy(yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")))
    target = document
    for key in mapping_path:
        target = target[key]
    if mutation == "unknown":
        target["unexpected_contract_key"] = 1
    else:
        del target[next(iter(target))]
    config_path = tmp_path / "key-inventory.yaml"
    config_path.write_text(yaml.safe_dump(document), encoding="utf-8")

    with pytest.raises(ProtocolConfigError, match="keys mismatch"):
        load_protocol_config(config_path)


@pytest.mark.parametrize("value", [True, 1.5, "2"])
def test_protocol_config_requires_exact_integer_stage_order(
    tmp_path: Path, value
) -> None:
    document = copy.deepcopy(yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")))
    document["stages"][1]["order"] = value
    path = tmp_path / "bad-order.yaml"
    path.write_text(yaml.safe_dump(document), encoding="utf-8")
    with pytest.raises(ProtocolConfigError, match="YAML integer"):
        load_protocol_config(path)


def test_partial_and_multiple_manual_intervals_are_lossless():
    multiple = parse_manual_interval_cell("708~723 / 740~754")
    assert [(item.start_s, item.end_s, item.complete) for item in multiple] == [
        (708.0, 723.0, True),
        (740.0, 754.0, True),
    ]
    partial = parse_manual_interval_cell("~736")
    assert partial[0].start_s is None
    assert partial[0].end_s == 736.0
    assert partial[0].qc_flags == ("manual_start_missing",)


@pytest.mark.skipif(not WORKBOOK_PATH.is_file(), reason="private workbook is not restored")
def test_spreadsheet_is_read_only_and_all_session_action_issue_metadata_is_structured():
    before = (WORKBOOK_PATH.stat().st_size, WORKBOOK_PATH.stat().st_mtime_ns)
    config, by_session = _loaded()
    after = (WORKBOOK_PATH.stat().st_size, WORKBOOK_PATH.stat().st_mtime_ns)
    assert after == before
    assert len(by_session) == 30

    s02 = by_session["S02_RJS"]
    assert s02.acquisition_date == "2026-07-20"
    assert s02.phase7_assignment == "Dodge"
    assert s02.manual_intervals["phase1"][0].start_s == 25.0
    assert s02.manual_intervals["phase2"][0].end_s == 611.0
    assert len(s02.manual_intervals["phase3"]) == 2
    assert "missing_stage_marker" in s02.qc_flags

    s03 = by_session["S03_PSJ"]
    assert s03.manual_intervals["phase3"][0].start_s is None
    assert s03.manual_intervals["phase3"][0].end_s == 736.0
    assert "excess_markers" in s03.qc_flags

    s04 = by_session["S04_KTW"]
    assert s04.fall_protocol_candidates[0].protocol_id == config.fall_v2_id
    assert s04.fall_protocol_candidates[0].confidence == pytest.approx(0.98)
    assert "retry_reported" in s04.qc_flags
    assert "phase7_may_be_unusable" in s04.qc_flags

    assert {
        record.phase7_assignment
        for record in by_session.values()
        if record.session_number <= 10
    } == {"Dodge"}
    assert {record.phase7_assignment for record in by_session.values()} == {
        "Dodge",
        "Strike",
        "Kick",
    }
    assert all(record.annotation_inference_feature_allowed is False for record in by_session.values())
    assert all(
        not (record.issue_text or record.extra_notes) or record.issues
        for record in by_session.values()
    )


@pytest.mark.skipif(not WORKBOOK_PATH.is_file(), reason="private workbook is not restored")
def test_manual_timing_has_priority_and_preserves_attempts_and_qc():
    config, by_session = _loaded()
    s02 = by_session["S02_RJS"]
    deliberately_conflicting = [
        BoundaryCandidate(100.0, 1.0, "radar_marker", "phase1", "start"),
        BoundaryCandidate(500.0, 1.0, "radar_marker", "phase1", "end"),
    ]
    decoded = decode_ordered_protocol(
        duration_s=1400.0,
        config=config,
        candidates=deliberately_conflicting,
        session_record=s02,
    )
    assert decoded.stage("phase1").start.time_s == 25.0
    assert decoded.stage("phase1").end.time_s == 217.0
    assert decoded.stage("phase1").start.status == "auto"
    assert decoded.stage("phase2").start.time_s == 245.0
    assert decoded.stage("phase2").end.time_s == 611.0
    assert len(decoded.stage("phase3").attempts) == 2
    assert {
        attempt.interpretation for attempt in decoded.stage("phase3").attempts
    } == {"manual_subinterval_not_necessarily_retry"}
    assert decoded.stage("phase4").status == "review"
    assert "missing_stage_marker" in decoded.stage("phase4").qc_flags
    assert decoded.phase7_assignment == "Dodge"


@pytest.mark.skipif(not WORKBOOK_PATH.is_file(), reason="private workbook is not restored")
def test_partial_manual_boundary_and_biopac_provenance_are_explicit():
    config, by_session = _loaded()
    s03 = by_session["S03_PSJ"]
    candidates = [
        BoundaryCandidate(
            690.0,
            1.0,
            "biopac_marker",
            "phase3",
            "start",
            biopac_derived=True,
        )
    ]
    decoded = decode_ordered_protocol(
        duration_s=1550.0,
        config=config,
        candidates=candidates,
        session_record=s03,
    )
    assert decoded.stage("phase3").start.time_s == 690.0
    assert decoded.stage("phase3").end.time_s == 736.0
    assert decoded.stage("phase3").status == "uncertain"
    assert decoded.used_biopac_derived_annotations is True
    assert decoded.annotation_inference_feature_allowed is False
    assert decoded.biopac_annotation_inference_feature_allowed is False
    assert decoded.to_dict()["annotation_contract"]["inference_feature_allowed"] is False


def test_ordered_dynamic_programming_and_window_transition_guard():
    config = load_protocol_config(CONFIG_PATH)
    nominal_boundaries = [
        25.0,
        205.0,
        250.0,
        610.0,
        685.0,
        740.0,
        830.0,
        980.0,
        1025.0,
        1195.0,
        1245.0,
        1270.0,
        1300.0,
        1315.0,
    ]
    candidates = []
    for stage_index, stage in enumerate(config.stages):
        candidates.extend(
            [
                BoundaryCandidate(
                    nominal_boundaries[2 * stage_index],
                    1.0,
                    "fused_marker",
                    stage.stage_id,
                    "start",
                ),
                BoundaryCandidate(
                    nominal_boundaries[2 * stage_index + 1],
                    1.0,
                    "fused_marker",
                    stage.stage_id,
                    "end",
                ),
            ]
        )
    decoded = decode_ordered_protocol(duration_s=1400.0, config=config, candidates=candidates)
    boundaries = [
        value
        for stage in decoded.stages
        for value in (stage.start.time_s, stage.end.time_s)
    ]
    assert boundaries == sorted(boundaries)
    assert boundaries == nominal_boundaries

    stable = assign_window_to_stage(100.0, 132.0, decoded)
    assert stable.stage_id == "phase1"
    assert not stable.eligible_for_stage_metrics
    assert not stable.transition_guard_triggered
    assert stable.reason == "stage_uncertain"

    auto_protocol = replace(
        decoded,
        stages=(replace(decoded.stages[0], status="auto"), *decoded.stages[1:]),
    )
    auto = assign_window_to_stage(100.0, 132.0, auto_protocol)
    assert auto.stage_id == "phase1"
    assert auto.eligible_for_stage_metrics
    assert auto.reason == "assigned"

    for status, reason in (
        ("review", "stage_requires_review"),
        ("unexpected", "stage_status_invalid"),
    ):
        changed = replace(
            decoded,
            stages=(replace(decoded.stages[0], status=status), *decoded.stages[1:]),
        )
        ineligible = assign_window_to_stage(100.0, 132.0, changed)
        assert ineligible.stage_id == "phase1"
        assert not ineligible.eligible_for_stage_metrics
        assert ineligible.reason == reason

    transition = assign_window_to_stage(190.0, 222.0, decoded)
    assert transition.stage_id is None
    assert transition.transition_guard_triggered
    assert not transition.eligible_for_stage_metrics
    assert transition.reason == "transition_guard"

    gap = assign_window_to_stage(212.0, 244.0, decoded)
    assert gap.stage_id is None
    assert not gap.eligible_for_stage_metrics
    assert gap.reason == "insufficient_stage_overlap"

    phase7 = assign_window_to_stage(1301.0, 1314.0, decoded)
    assert phase7.stage_id is None  # within the 2 s transition guards
