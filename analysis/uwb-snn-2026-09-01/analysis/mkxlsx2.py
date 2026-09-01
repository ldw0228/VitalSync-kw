# -*- coding: utf-8 -*-
import numpy as np, json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
W='_work'
SY={o['s']:o for o in json.load(open('sync12_final.json'))}
XL=json.load(open('xl_exp2.json')); LM=json.load(open('landmarks.json'))
AB=json.load(open('ab2.json')); TR=json.load(open('train_v2.json'))
SCRIPT=135.0
FONT='Arial'
HDR=PatternFill('solid',fgColor='DDE5EC'); BAD=PatternFill('solid',fgColor='FCE4E0')
OK=PatternFill('solid',fgColor='E8F4EC'); THIN=Side(style='thin',color='BFC9D0')
BD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
def sheet(wb,name,cols,rows,widths=None,fills=None):
    ws=wb.create_sheet(name)
    for j,c in enumerate(cols,1):
        cell=ws.cell(1,j,c); cell.font=Font(name=FONT,bold=True,size=10)
        cell.fill=HDR; cell.border=BD; cell.alignment=Alignment(horizontal='center',wrap_text=True)
    for i,r in enumerate(rows,2):
        for j,v in enumerate(r,1):
            cell=ws.cell(i,j,v); cell.font=Font(name=FONT,size=10); cell.border=BD
            if isinstance(v,(int,float)): cell.alignment=Alignment(horizontal='center')
        if fills:
            f=fills(r)
            if f:
                for j in range(1,len(cols)+1): ws.cell(i,j).fill=f
    for j,c in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(j)].width=(widths[j-1] if widths else 14)
    ws.freeze_panes='A2'
    return ws
def markers(rsp,fsb,thr=8.5,merge=4.0):
    ab=rsp>thr; d=np.diff(np.r_[0,ab.astype(int),0])
    on=np.where(d==1)[0]; off=np.where(d==-1)[0]-1
    mk=[(a+int(np.argmax(rsp[a:b+1])))/fsb for a,b in zip(on,off)]
    if not mk: return []
    mk=sorted(mk); out=[]; grp=[mk[0]]
    for x in mk[1:]:
        if x-grp[-1]<=merge: grp.append(x)
        else: out.append(float(np.mean(grp))); grp=[x]
    out.append(float(np.mean(grp))); return out
def start(v):
    if not v or '?' in str(v): return None
    try: return float(str(v).split('~')[0])
    except Exception: return None
B2=[('평소 호흡',-140,-78),('느린 호흡',-78,-16),('무호흡',-16,16),
    ('회복 호흡',16,78),('운동',94,156),('운동 후 호흡',172,234)]
permae={}
for tag,key in [('마커','marker'),('지형','landmark')]:
    for x in AB[key]: permae.setdefault(x['s'],{}).setdefault(tag,[]).append(x['err'])
wb=openpyxl.Workbook(); wb.remove(wb.active)

# 1) 구간표
cols=['피험자','마커 수','실험2 시작(마커기반)','무호흡 중심(마커기반)','무호흡 중심(지형지물)',
      '두 방식 차이(초)','offset(BIOPAC-레이더)','실험1 블록1','실험1 블록2','실험1 블록3']
cols+=['%s (지형지물)'%n for n,_,_ in B2]
cols+=['마커차 실1시작','마커차 실1종료','마커차 실2시작','호흡수오차 마커기반','호흡수오차 지형지물','판정']
rows=[]
for s,o in sorted(SY.items()):
    num=int(s[1:3])
    xr=start(XL.get(str(num),{}).get('rad'))
    apm=xr+SCRIPT if xr is not None else None
    apl=o['ap'] if o['grade']!='실패' else None
    mk=[]
    p=os.path.join(W,s+'_mot.npz')
    if os.path.exists(p):
        z=np.load(p); mk=markers(z['rsp'].astype(float),float(z['fsb']))
    r=[s,len(mk),round(xr,1) if xr is not None else '—',
       round(apm,1) if apm else '—', round(apl,1) if apl else '—',
       round(apm-apl,1) if (apm and apl) else '—', round(o['off'],1) if apl else '—']
    if apl and s in LM:
        for a,b in LM[s]['blocks']: r.append('%.1f ~ %.1f'%(a,b))
        for n,a0,a1 in B2: r.append('%.1f ~ %.1f'%(apl+a0,apl+a1))
        near=lambda t: min(mk,key=lambda m:abs(m-t)) if mk else float('nan')
        bl=LM[s]['blocks']; off=o['off']
        d=[round((bl[0][0]+off)-near(bl[0][0]+off),1),round((bl[2][1]+off)-near(bl[2][1]+off),1),
           round((apl-140+off)-near(apl-140+off),1)]
        r+=d
    else:
        r+=['—']*3+['—']*6+['—']*3
    pm=permae.get(s,{}).get('마커'); pl=permae.get(s,{}).get('지형')
    r.append(round(float(np.mean(pm)),2) if pm else '—')
    r.append(round(float(np.mean(pl)),2) if pl else '—')
    if apl is None: v='실험1 미검출 — 제외'
    elif apm is None: v='마커 경계 미검출 (지형지물만 사용)'
    elif isinstance(r[13],float) and max(abs(r[11]),abs(r[12]),abs(r[13]))>30: v='마커와 불일치 — 확인 필요'
    else: v='정상'
    r.append(v); rows.append(r)
def fill1(r):
    v=r[-1]
    if v.startswith('실험1') or v.startswith('마커와'): return BAD
    if v=='정상': return None
    return None
sheet(wb,'1_구간표',cols,rows,[13,8]+[16]*5+[17]*3+[17]*6+[13]*3+[15,15,26],fill1)

# 2) 모델 결과
cols2=['자르는 기준','모델','인코딩','호흡수 오차 MAE','1회/분 이내','2.5회/분 이내',
       '고전 기준선 MAE','고전 대비 개선','파라미터','은닉층 발화율','조기종료 에폭 (4폴드)',
       '학습 시간(초)','피험자','창 수']
MODEN={'marker':'마커 기반','landmark':'지형지물 기반'}
rows2=[]
for mode in ['marker','landmark']:
    sub=[r for r in TR if r['mode']==mode]
    b=sub[0]
    rows2.append([MODEN[mode],'고전 신호처리','—',round(b['base_mae'],3),round(b['base_p1'],3),'—',
                  round(b['base_mae'],3),0.0,0,'—','—',0,b['subj'],b['n']])
    for r in sub:
        rows2.append([MODEN[mode],r['kind'],r['enc'],
            round(r['mae'],3),round(r['p1'],3),round(r['p25'],3),round(r['base_mae'],3),
            round(r['base_mae']-r['mae'],3),r['params'],round(r['spike'],3) if r['spike'] else '—',
            ', '.join(str(e) for e in r['stop_epochs']),
            r['secs'],r['subj'],r['n']])
BESTMAE=min(r['mae'] for r in TR)
def fill2(r):
    if r[1]=='고전 신호처리': return HDR
    if abs(r[3]-round(BESTMAE,3))<1e-9: return OK
    return None
sheet(wb,'2_모델결과',cols2,rows2,[13,13,12,15,13,14,15,13,11,13,18,12,8,8],fill2)

# 3) 학습 설정
cols3=['항목','값','비고']
rows3=[['과제','9개 후보 중 신뢰할 추정 고르기','후보 = 레이더 3대 × 복조 3방식'],
 ['입력','후보별 호흡 파형 100 step (30초 @3.33Hz)','창당 후보 9개'],
 ['보조 입력','신뢰도 특징 10종','배음비 h2·h3, 기본파비중, 스펙트럼 엔트로피, 2등봉우리비, 전후 드리프트, 표준편차, 적합잔차, 레이더번호, 복조번호'],
 ['','+ 후보 관계 특징 3종','추정값, 중앙값과의 차, 합의수(±1 BPM 이내 동의 후보 수)'],
 ['출력','후보별 신뢰도 점수 1개','가장 낮은 점수의 후보를 채택'],
 ['손실','MSE(log(1+오차)) + 0.5 × CrossEntropy(최선후보)','회귀 + 순위 결합'],
 ['SNN 구조','Linear→LIF(96) → Linear→LIF(48) → head(48+13→48→1)','은닉 2층'],
 ['ANN 구조','Linear→ReLU(96) → Linear→ReLU(48) → head(48+13→48→1)','SNN과 층·폭 동일'],
 ['LIF 뉴런','snn.Leaky, β 초기 0.9 (학습 가능)','surrogate gradient = arctangent'],
 ['시간 전개','100 step','SNN만 해당'],
 ['에폭','상한 60, 조기종료 (인내 10)','검증셋 MAE가 10에폭 동안 나아지지 않으면 중단'],
 ['분할','피험자 단위 학습/검증/평가 3분할','평가 폴드 f, 검증 폴드 (f+1)%4, 나머지 학습'],
 ['배치','32',''],
 ['학습률','2e-3','Adam, weight decay 1e-4'],
 ['그래디언트 클리핑','norm 5.0',''],
 ['교차검증','피험자 분리 4-fold','모든 피험자가 한 번씩 평가에 포함'],
 ['시드','0 고정','torch/numpy 동일'],
 ['모델 선택','검증셋 최고 에폭의 가중치 복원','평가셋은 마지막에 한 번만 사용'],
 ['인코딩 5종','direct / rate / delta / step-forward / population','direct는 스파이크 아님(아날로그 주입)'],
 ['창','30초, 5초 이동','겹치는 창 — 피험자 단위로 분할해 누수 방지'],
 ['배경 제거','되먹임 필터 β=0.98','빈 방 녹화 없이 정적 배경 제거'],
]
sheet(wb,'3_학습설정',cols3,rows3,[20,45,60])

# 4) 요약
EM=np.array([x['err'] for x in AB['marker'] if x['s']!='S01_CMS'])
EL=np.array([x['err'] for x in AB['landmark']])
sm=set(x['s'] for x in AB['marker'])-{'S01_CMS'}; sl=set(x['s'] for x in AB['landmark'])
com=sm&sl
em=np.array([x['err'] for x in AB['marker'] if x['s'] in com])
el=np.array([x['err'] for x in AB['landmark'] if x['s'] in com])
best={m:min([r for r in TR if r['mode']==m],key=lambda r:r['mae']) for m in ['marker','landmark']}
cols4=['항목','마커 기반','지형지물 기반']
rows4=[['사용 피험자',len(sm),len(sl)],
 ['창 수',int(len(EM)),int(len(EL))],
 ['고전 기준선 MAE',round(float(EM.mean()),2),round(float(EL.mean()),2)],
 ['고전 1회/분 이내',round(float(np.mean(EM<=1)),3),round(float(np.mean(EL<=1)),3)],
 ['','',''],
 ['같은 %d명만 비교 — 고전 MAE'%len(com),round(float(em.mean()),2),round(float(el.mean()),2)],
 ['같은 %d명만 비교 — 1회/분 이내'%len(com),round(float(np.mean(em<=1)),3),round(float(np.mean(el<=1)),3)],
 ['','',''],
 ['최고 모델',best['marker']['kind']+' '+best['marker']['enc'],best['landmark']['kind']+' '+best['landmark']['enc']],
 ['최고 모델 MAE',round(best['marker']['mae'],2),round(best['landmark']['mae'],2)],
 ['최고 모델 1회/분 이내',round(best['marker']['p1'],3),round(best['landmark']['p1'],3)],
 ['ANN MAE',round([r for r in TR if r['mode']=='marker' and r['kind']=='ANN'][0]['mae'],2),
            round([r for r in TR if r['mode']=='landmark' and r['kind']=='ANN'][0]['mae'],2)],
 ['ANN 파라미터',17377,17377],
 ['최고 SNN 파라미터',best['marker']['params'],best['landmark']['params']],
 ['','',''],
 ['평가 방식','피험자 분리 4-fold · 학습/검증/평가 3분할','좌우 동일'],
 ['에폭 결정','검증셋 조기종료 (상한 60, 인내 10)','평가셋은 마지막 1회만'],
]
sheet(wb,'4_요약',cols4,rows4,[34,18,18])
# 5) 판단 근거
cols5=['설정 / 결정','값','왜 그렇게 정했나','출처']
rows5=[
 ['창 길이','30초','호흡 4.8회/분(12.5초 주기)에서도 주기 2회 이상이 들어가야 스펙트럼 봉우리가 선다','우리 판단 · 문헌 관행 20~60초'],
 ['창 이동','5초','창 수를 늘리되 인접 창이 과도하게 겹치지 않는 절충. 피험자 단위 분할로 누수는 차단','우리 판단'],
 ['재표본 주파수','10 Hz → 모델 입력 3.33 Hz','호흡 대역 0.08~0.6 Hz의 5배 이상. 40 fps 원본은 불필요하게 무거움','Nyquist'],
 ['배경 제거 β','0.98','논문 값 0.9는 10 Hz에서 차단 0.16 Hz — 호흡 대역을 잘라먹음. 0.98은 0.032 Hz','MoRe-Fi 식, 값은 우리가 재계산'],
 ['가슴 bin 선택','창 안에서 호흡 대역 에너지 최대 bin','피험자가 움직이므로 고정 bin은 성립하지 않음','우리 판단'],
 ['후보 9개','레이더 3대 × 복조 3방식','최선 레이더가 1번 38회·2번 36회·3번 38회로 균등 — 고정 규칙 불가','우리 데이터로 확인'],
 ['4-fold','피험자 단위','같은 사람의 창이 학습과 평가에 함께 들어가면 누수. 28명이라 5-fold 이상은 폴드가 너무 작음','표준'],
 ['3분할','학습/검증/평가','에폭을 평가셋으로 고르면 결과가 부풀려짐 (최대 0.29 확인)','표준 · 교정 사항'],
 ['조기종료','상한 60, 인내 10','에폭 수를 사람이 정하지 않고 데이터가 정하게','표준'],
 ['학습률 2e-3','Adam','Adam 기본값 1e-3의 2배 — 60에폭 상한에 맞춰 수렴 속도 확보','관행 범위'],
 ['배치 32','','작은 데이터셋(창 588~784)에서 관행 범위','관행'],
 ['weight decay 1e-4','','과적합 억제 관행값','관행'],
 ['그래디언트 클리핑','norm 5.0','SNN 시간 전개에서 그래디언트 폭발 방지','SNN 관행'],
 ['LIF β','0.9 초기, 학습 가능','snntorch 예제 기본값에서 출발, 학습으로 조정되게 둠','snntorch'],
 ['surrogate','arctangent','snn.Leaky 표준 선택지','snntorch'],
 ['손실','MSE(log1p 오차) + 0.5·CE(최선후보)','오차 크기(회귀)와 어느 후보가 최선인지(순위)를 함께 학습','우리 설계'],
 ['은닉층 96→48','','**탐색하지 않음.** 데이터 규모 보고 임의로 정함 — 근거 없음','미해결'],
 ['보조 특징 13종','','**어느 것이 기여하는지 미확인** — 절제 실험은 교정 전 방법론','미해결'],
 ['손실 가중 0.5','','**탐색하지 않음**','미해결'],
]
def fill5(r):
    return BAD if r[3]=='미해결' else None
sheet(wb,'5_판단근거',cols5,rows5,[22,26,62,26],fill5)

wb.save('실험비교_마커vs지형지물.xlsx')
print('저장 완료')
